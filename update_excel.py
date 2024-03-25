import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import PatternFill


class ExcelHyperlinkUpdater:
    def __init__(self, filepath, hex_color):
        self.filepath = filepath
        self.hex_color = hex_color.lstrip('#')

    def copy_and_update_excel(self, data_dict, new_filepath):
        # Copy the original Excel file to a new location
        shutil.copy(self.filepath, new_filepath)

        # Now, work with the copied file
        wb = load_workbook(new_filepath)
        ws = wb.create_sheet("Hyperlinks", 0)

        # Define the fill color for light red
        fill = PatternFill(start_color=self.hex_color, end_color=self.hex_color, fill_type="solid")

        # Header row
        ws['A1'] = 'Cell Data'
        ws['B1'] = 'PDF page number'
        ws['C1'] = 'Tab name'

        # Populate the worksheet with data, hyperlinks, and tab names
        row_num = 2
        for tab_name, cells in data_dict.items():
            for cell_address, info in cells.items():
                cell_data = info['Cell_data']
                page_number = info['Page_number']

                ws[f'A{row_num}'] = cell_data
                ws[f'B{row_num}'] = page_number
                ws[f'C{row_num}'] = tab_name

                # Format the sheet name properly for hyperlink
                formatted_sheet_name = f"'{tab_name}'" if ' ' in tab_name or '!' in tab_name or ',' in tab_name \
                    else tab_name

                # Create a hyperlink to the original cell location in the respective sheet
                hyperlink_target = f"#{formatted_sheet_name}!{cell_address}"
                ws[f'A{row_num}'].hyperlink = Hyperlink(ref="", target=hyperlink_target,
                                                        tooltip=f"Go to {cell_address} in {tab_name}")

                # Coloring the destination row in light red
                destination_sheet = wb[tab_name]
                destination_row = int(cell_address[1:])
                for cell in destination_sheet[destination_row]:
                    cell.fill = fill

                row_num += 1

        # Save the workbook
        wb.save(new_filepath)
