import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelProcessor:
    def __init__(self, filepath, header):
        self.filepath = filepath
        self.header = header
        self.result_dict = {}

    def process_excel(self):
        # Open the workbook with openpyxl to work with cell addresses
        wb = load_workbook(self.filepath)

        # Iterate over each sheet in the Excel file
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            df = pd.read_excel(self.filepath, sheet_name=sheet_name)

            # Attempt to find the column index
            id_column_index = None
            for column in df.columns:
                if column.upper() == self.header.upper():  # Case-insensitive comparison
                    id_column_index = df.columns.get_loc(column)
                    break

            # If column exists
            if id_column_index is not None:
                sheet_dict = {}
                # Starting from row 2 to skip header
                for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=id_column_index+1, values_only=True), start=2):
                    cell_value = row[id_column_index]
                    if cell_value:  # Check if cell is not empty
                        cell_address = f"{get_column_letter(id_column_index + 1)}{idx}"
                        sheet_dict[cell_address] = cell_value

                # Add the sheet dictionary to the result dictionary
                self.result_dict[sheet_name] = sheet_dict

        wb.close()
        return self.result_dict

